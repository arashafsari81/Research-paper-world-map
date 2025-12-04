from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from typing import Optional
from datetime import datetime
from csv_processor import CSVProcessor
from openpyxl import Workbook
from io import BytesIO
import shutil


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Global data cache - stores data for each year filter
cached_data = {}  # year -> processed_data
cached_stats = {}  # year -> stats

def load_data(year_filter=None):
    """Load and process CSV data with optional year filter (single year or range)."""
    global cached_data, cached_stats
    
    # Create cache key that handles both single year and range
    if isinstance(year_filter, (tuple, list)):
        cache_key = f"{year_filter[0]}-{year_filter[1]}"
    else:
        cache_key = year_filter if year_filter else 'all'
    
    # Return cached data if available
    if cache_key in cached_data:
        return cached_data[cache_key], cached_stats[cache_key]
    
    # Priority: uploaded file > default uncleaned file > old cleaned file
    csv_path = '/app/uploaded_dataset.csv'
    if not os.path.exists(csv_path):
        csv_path = '/app/Scopus_Data_APU_2021_Dec_2025_Complete.csv'
    if not os.path.exists(csv_path):
        csv_path = '/app/APU_publications_2021_2025_cleaned_Final.csv'
    if not os.path.exists(csv_path):
        logger.error(f"CSV file not found at {csv_path}")
        return None, None
    
    try:
        processor = CSVProcessor(csv_path, year_filter=year_filter)
        processor.load_csv().process_data()
        data = processor.get_processed_data()
        stats = processor.get_stats()
        
        # Cache the results
        cached_data[cache_key] = data
        cached_stats[cache_key] = stats
        
        logger.info(f"Data loaded successfully for year={year_filter}: {stats}")
        return data, stats
    except Exception as e:
        logger.error(f"Error processing CSV: {e}")
        import traceback
        traceback.print_exc()
        return None, None


# Routes
@api_router.get("/")
async def root():
    return {"message": "Research Papers World Map API"}

@api_router.post("/upload-dataset")
async def upload_dataset(file: UploadFile = File(...)):
    """Upload a new CSV dataset to replace the current one."""
    global cached_data, cached_stats
    
    # Validate file type
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Only CSV files are allowed")
    
    try:
        # Save the uploaded file
        upload_path = '/app/uploaded_dataset.csv'
        with open(upload_path, 'wb') as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        # Clear the cache to force reload with new data
        cached_data.clear()
        cached_stats.clear()
        
        # Try to load and validate the new dataset
        try:
            data, stats = load_data()
            if data is None or stats is None:
                # Remove the invalid file
                os.remove(upload_path)
                raise HTTPException(status_code=400, detail="Invalid CSV format. Please ensure the CSV has the required columns.")
            
            logger.info(f"New dataset uploaded successfully: {stats}")
            return {
                "message": "Dataset uploaded successfully",
                "stats": stats,
                "filename": file.filename
            }
        except Exception as e:
            # Remove the invalid file
            if os.path.exists(upload_path):
                os.remove(upload_path)
            raise HTTPException(status_code=400, detail=f"Error processing CSV: {str(e)}")
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error uploading dataset: {e}")
        raise HTTPException(status_code=500, detail=f"Error uploading file: {str(e)}")

@api_router.get("/current-dataset-info")
async def get_current_dataset_info():
    """Get information about the currently loaded dataset."""
    # Check which file is being used
    if os.path.exists('/app/uploaded_dataset.csv'):
        dataset_source = "uploaded"
        file_path = '/app/uploaded_dataset.csv'
    elif os.path.exists('/app/Scopus_Data_APU_2021_Dec_2025_Complete.csv'):
        dataset_source = "default"
        file_path = '/app/Scopus_Data_APU_2021_Dec_2025_Complete.csv'
    else:
        dataset_source = "legacy"
        file_path = '/app/APU_publications_2021_2025_cleaned_Final.csv'
    
    # Get file stats
    file_stat = os.stat(file_path)
    file_size_mb = file_stat.st_size / (1024 * 1024)
    
    return {
        "source": dataset_source,
        "file_path": file_path,
        "file_size_mb": round(file_size_mb, 2),
        "last_modified": datetime.fromtimestamp(file_stat.st_mtime).isoformat()
    }

@api_router.get("/stats")
async def get_stats(year: Optional[int] = None, start_year: Optional[int] = None, end_year: Optional[int] = None):
    """Get overall statistics with optional year filter (single year or range)."""
    # Determine year filter type
    if start_year and end_year:
        year_filter = (start_year, end_year)
    elif year:
        year_filter = year
    else:
        year_filter = None
    
    data, stats = load_data(year_filter=year_filter)
    
    if stats is None:
        raise HTTPException(status_code=500, detail="Data not loaded")
    
    # Count only countries with valid coordinates (that appear on map)
    countries_on_map = 0
    if data:
        for country in data:
            if country['lat'] != 0 or country['lng'] != 0:
                countries_on_map += 1
    
    return {
        **stats,
        'totalCountries': countries_on_map  # Only show countries visible on map
    }

@api_router.get("/data/countries")
async def get_countries(year: Optional[int] = None, start_year: Optional[int] = None, end_year: Optional[int] = None):
    """Get all countries with paper counts and coordinates, with optional year filter."""
    # Determine year filter type
    if start_year and end_year:
        year_filter = (start_year, end_year)
    elif year:
        year_filter = year
    else:
        year_filter = None
    
    data, stats = load_data(year_filter=year_filter)
    
    if data is None:
        raise HTTPException(status_code=500, detail="Data not loaded")
    
    # Return simplified country data for map
    countries = []
    for country in data:
        countries.append({
            'id': country['id'],
            'name': country['name'],
            'lat': country['lat'],
            'lng': country['lng'],
            'paperCount': country['paperCount']
        })
    
    return {'countries': countries}

@api_router.get("/data/country/{country_id}")
async def get_country(country_id: str, year: Optional[int] = None, start_year: Optional[int] = None, end_year: Optional[int] = None):
    """Get universities for a specific country with optional year filter."""
    # Determine year filter type
    if start_year and end_year:
        year_filter = (start_year, end_year)
    elif year:
        year_filter = year
    else:
        year_filter = None
    
    data, stats = load_data(year_filter=year_filter)
    
    if data is None:
        raise HTTPException(status_code=500, detail="Data not loaded")
    
    # Find country
    country = None
    for c in data:
        if c['id'] == country_id:
            country = c
            break
    
    if not country:
        raise HTTPException(status_code=404, detail="Country not found")
    
    # Return country with simplified universities (without full author data)
    universities = []
    for uni in country['universities']:
        universities.append({
            'id': uni['id'],
            'name': uni['name'],
            'paperCount': uni['paperCount'],
            'citationCount': uni.get('citationCount', 0),
            'authors': len(uni['authors'])
        })
    
    return {
        'country': {
            'id': country['id'],
            'name': country['name'],
            'paperCount': country['paperCount'],
            'citationCount': country.get('citationCount', 0),
            'universities': universities
        }
    }

@api_router.get("/data/university/{country_id}/{university_id}")
async def get_university(country_id: str, university_id: str, year: Optional[int] = None):
    """Get authors for a specific university with optional year filter."""
    data, stats = load_data(year_filter=year)
    
    if data is None:
        raise HTTPException(status_code=500, detail="Data not loaded")
    
    # Find country and university
    country = None
    for c in data:
        if c['id'] == country_id:
            country = c
            break
    
    if not country:
        raise HTTPException(status_code=404, detail="Country not found")
    
    university = None
    for uni in country['universities']:
        if uni['id'] == university_id:
            university = uni
            break
    
    if not university:
        raise HTTPException(status_code=404, detail="University not found")
    
    # Return university with simplified authors (without papers)
    authors = []
    for author in university['authors']:
        authors.append({
            'id': author['id'],
            'name': author['name'],
            'affiliation': author['affiliation'],
            'paperCount': author['paperCount'],
            'citationCount': author.get('citationCount', 0)
        })
    
    return {
        'university': {
            'id': university['id'],
            'name': university['name'],
            'country': country['name'],
            'paperCount': university['paperCount'],
            'citationCount': university.get('citationCount', 0),
            'authors': authors
        }
    }

@api_router.get("/data/author/{country_id}/{university_id}/{author_id}")
async def get_author(country_id: str, university_id: str, author_id: str, year: Optional[int] = None):
    """Get papers for a specific author with optional year filter."""
    data, stats = load_data(year_filter=year)
    
    if data is None:
        raise HTTPException(status_code=500, detail="Data not loaded")
    
    # Find country, university, and author
    country = None
    for c in data:
        if c['id'] == country_id:
            country = c
            break
    
    if not country:
        raise HTTPException(status_code=404, detail="Country not found")
    
    university = None
    for uni in country['universities']:
        if uni['id'] == university_id:
            university = uni
            break
    
    if not university:
        raise HTTPException(status_code=404, detail="University not found")
    
    author = None
    for a in university['authors']:
        if a['id'] == author_id:
            author = a
            break
    
    if not author:
        raise HTTPException(status_code=404, detail="Author not found")
    
    return {'author': author}

@api_router.get("/search")
async def search(q: Optional[str] = None, year: Optional[int] = None):
    """Search across all data."""
    data, stats = load_data(year_filter=year)
    
    if data is None:
        raise HTTPException(status_code=500, detail="Data not loaded")
    
    # Note: Search filtering should be done on frontend for better performance
    # This endpoint returns all data, frontend will filter
    return {'countries': data}

@api_router.get("/export/papers")
async def export_papers(year: Optional[int] = None):
    """Export all paper titles to Excel."""
    data, stats = load_data(year_filter=year)
    
    if data is None:
        raise HTTPException(status_code=500, detail="Data not loaded")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Papers"
    
    # Headers
    ws.append(["#", "Title", "Year", "Source", "Citations", "DOI", "Authors"])
    
    # Collect all unique papers
    papers_dict = {}
    for country in data:
        for uni in country['universities']:
            for author in uni['authors']:
                for paper in author['papers']:
                    if paper['id'] not in papers_dict:
                        papers_dict[paper['id']] = paper
    
    # Add papers
    for idx, (paper_id, paper) in enumerate(papers_dict.items(), 1):
        ws.append([
            idx,
            paper['title'],
            paper['year'],
            paper['source'],
            paper.get('cited_by', 0),
            paper.get('doi', ''),
            ', '.join(paper.get('authors', []))
        ])
    
    # Adjust column widths
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 50
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"papers_export_{year if year else 'all_years'}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/export/authors")
async def export_authors(year: Optional[int] = None):
    """Export all authors to Excel."""
    data, stats = load_data(year_filter=year)
    
    if data is None:
        raise HTTPException(status_code=500, detail="Data not loaded")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Authors"
    
    # Headers
    ws.append(["#", "Author Name", "Author ID", "Affiliation", "Country", "Paper Count"])
    
    # Collect all unique authors
    authors_dict = {}
    for country in data:
        for uni in country['universities']:
            for author in uni['authors']:
                if author['id'] not in authors_dict:
                    authors_dict[author['id']] = {
                        'name': author['name'],
                        'id': author['id'],
                        'affiliation': author['affiliation'],
                        'country': country['name'],
                        'paperCount': author['paperCount']
                    }
    
    # Add authors
    for idx, (author_id, author) in enumerate(authors_dict.items(), 1):
        ws.append([
            idx,
            author['name'],
            author['id'],
            author['affiliation'],
            author['country'],
            author['paperCount']
        ])
    
    # Adjust column widths
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 25
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"authors_export_{year if year else 'all_years'}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/export/countries")
async def export_countries(year: Optional[int] = None):
    """Export all countries to Excel."""
    data, stats = load_data(year_filter=year)
    
    if data is None:
        raise HTTPException(status_code=500, detail="Data not loaded")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Countries"
    
    # Headers
    ws.append(["#", "Country Name", "Paper Count", "Universities Count", "Latitude", "Longitude"])
    
    # Add countries
    for idx, country in enumerate(data, 1):
        ws.append([
            idx,
            country['name'],
            country['paperCount'],
            len(country['universities']),
            country['lat'],
            country['lng']
        ])
    
    # Adjust column widths
    ws.column_dimensions['B'].width = 30
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"countries_export_{year if year else 'all_years'}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/export/universities")
async def export_universities(year: Optional[int] = None):
    """Export all universities to Excel."""
    data, stats = load_data(year_filter=year)
    
    if data is None:
        raise HTTPException(status_code=500, detail="Data not loaded")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Universities"
    
    # Headers
    ws.append(["#", "University Name", "Country", "Paper Count", "Authors Count"])
    
    # Collect all universities
    universities_list = []
    for country in data:
        for uni in country['universities']:
            universities_list.append({
                'name': uni['name'],
                'country': country['name'],
                'paperCount': uni['paperCount'],
                'authorsCount': len(uni['authors'])
            })
    
    # Sort by paper count
    universities_list.sort(key=lambda x: x['paperCount'], reverse=True)
    
    # Add universities
    for idx, uni in enumerate(universities_list, 1):
        ws.append([
            idx,
            uni['name'],
            uni['country'],
            uni['paperCount'],
            uni['authorsCount']
        ])
    
    # Adjust column widths
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 25
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"universities_export_{year if year else 'all_years'}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()